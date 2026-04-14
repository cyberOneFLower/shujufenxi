#!/usr/bin/env python3
"""生成《加密货币套利监控系统》功能点清单 Excel。"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "加密货币套利监控系统-功能点清单.xlsx"

HEAD = ("模块", "编号", "功能点", "说明/验收要点")

ROWS = [
    ("一、数据抓取与底层引擎", "F1.1", "多交易所 WebSocket 对接", "接入 Bitget、Gate、MEXC、OKX 等，实时订阅深度/盘口推送。"),
    ("一、数据抓取与底层引擎", "F1.2", "消息队列（MQ）削峰", "引入 RabbitMQ；采集先入队，消费端按能力拉取；极端行情防宕机、错乱、长延迟。"),
    ("一、数据抓取与底层引擎", "F1.3", "同秒时间戳对齐", "跨所价差基于同一秒内时间戳对齐，剔除时间差导致的假价差。"),
    ("一、数据抓取与底层引擎", "F1.4", "数据清洗与标准化", "各所格式统一为可计算结构（Bid/Ask、数量、时间戳等）。"),
    ("二、价差行情监控模块", "F2.1", "第一档盘口比对", "用各所买一 Bid1、卖一 Ask1 计算跨所价差百分比。"),
    ("二、价差行情监控模块", "F2.2", "五档盘口扩展", "买一至买五、卖一至卖五；用于均价等扩展计算（与实现口径一致）。"),
    ("二、价差行情监控模块", "F2.3", "买入/卖出总价", "总价 = 第一档价格 × 第一档数量。"),
    ("二、价差行情监控模块", "F2.4", "总价阈值过滤", "用户可设最小总价（如 N 美金）；低于阈值不展示。"),
    ("二、价差行情监控模块", "F2.5", "总价四档颜色", "≥1000 红；500–999 黄；100–499 蓝；<100 白/默认。"),
    ("二、价差行情监控模块", "F2.6", "列表排序", "支持按价差等维度排序。"),
    ("二、价差行情监控模块", "F2.7", "屏蔽/移除与黑名单", "一键移除交易对；写入过滤黑名单，后续不再展示。"),
    ("三、极端行情（暴涨暴跌）", "F3.1", "单币种 5 分钟滚动涨跌幅", "各交易所单币种，滚动 5 分钟窗口内计算涨跌幅。"),
    ("三、极端行情（暴涨暴跌）", "F3.2", "页面刷新频率", "3 秒或 5 秒自动刷新（可配置）。"),
    ("三、极端行情（暴涨暴跌）", "F3.3", "声音报警", "用户设涨跌幅阈值；触发时浏览器明显提示音。"),
    ("三、极端行情（暴涨暴跌）", "F3.4", "过滤与黑名单", "可永久屏蔽指定币种或平台。"),
    ("四、用户管理与系统管控", "F4.1", "多账号", "多子账号；独立密码、备注。"),
    ("四、用户管理与系统管控", "F4.2", "模块权限", "控制子账号是否使用暴涨暴跌模块（启用/停用）。"),
    ("四、用户管理与系统管控", "F4.3", "过滤列表（黑名单）独立页", "展示已屏蔽的价差对与暴涨暴跌币种；支持误删恢复。"),
    ("四、用户管理与系统管控", "F4.4", "紧急运维（说明）", "严重卡死时云主机重启；服务随主机自启；MQ/时序库等降低卡死概率。"),
    ("五、技术架构与数据流（非功能）", "T1", "Redis", "同秒对齐缓存、黑名单等高速读写。"),
    ("五、技术架构与数据流（非功能）", "T2", "时序数据库 TSDB", "5 分钟滚动与 Tick 追溯，避免 MySQL 写爆。"),
    ("五、技术架构与数据流（非功能）", "T3", "关系型库 MySQL", "用户、权限、黑名单配置、日志等低频持久化。"),
    ("五、技术架构与数据流（非功能）", "T4", "数据流", "采集→MQ→价差/异常服务→Redis/TSDB/MySQL→API/SSE/WS→前端。"),
    ("五、技术架构与数据流（非功能）", "T5", "交付物", "全部源码、部署文档；可选压力测试与验收。"),
    ("六、里程碑验收要点", "M1", "阶段1 Day1–6", "MQ 部署、≥3 家 WS、同秒对齐跑通、极端测试不宕机。"),
    ("六、里程碑验收要点", "M2", "阶段2 Day7–12", "价差前端、总价公式、阈值与四色、排序与移除。"),
    ("六、里程碑验收要点", "M3", "阶段3 Day13–16", "暴涨暴跌页、5 分钟逻辑、3–5s 刷新、声音报警、过滤列表。"),
    ("六、里程碑验收要点", "M4", "阶段4 Day17–20", "多账号与权限、联调压测、源码与文档移交。"),
    ("七、报价模块摘要（方案原文）", "B1", "底层引擎", "多线程/异步接入、MQ 削峰、同秒对齐。"),
    ("七、报价模块摘要（方案原文）", "B2", "价差模块", "一档数据、总价计算、四色、阈值隐藏。"),
    ("七、报价模块摘要（方案原文）", "B3", "暴涨暴跌", "5 分钟窗口、约 3 秒级刷新、Web 音频报警。"),
    ("七、报价模块摘要（方案原文）", "B4", "控制台", "黑名单打通、多用户隔离与权限、单点登录（方案报价条目中提及）。"),
    ("七、报价模块摘要（方案原文）", "B5", "测试部署", "压测、生产搭建、无加密源码与文档。"),
]


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "功能点清单"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="FFD9E1F2")
    for col, h in enumerate(HEAD, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)

    for i, row in enumerate(ROWS, start=2):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    widths = (28, 10, 26, 62)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    meta = wb.create_sheet("说明", 1)
    meta["A1"] = "文档来源"
    meta["B1"] = "《加密货币套利监控系统 - 开发方案》v1.0（2026.04.03）"
    meta["A2"] = "生成方式"
    meta["B2"] = "由项目脚本 scripts/generate_feature_excel.py 根据方案整理"
    meta["A3"] = "说明"
    meta["B3"] = "本表为需求/功能点清单，便于评审与验收；与当前代码实现可能有差异。"
    for r in range(1, 4):
        meta[f"B{r}"].alignment = Alignment(wrap_text=True)
    meta.column_dimensions["A"].width = 14
    meta.column_dimensions["B"].width = 70

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
